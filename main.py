import openpyxl


class Workbook:
    def __init__(self, file_name, out_name_and_ext=None):
        self.file_name = file_name
        self.workbook = openpyxl.load_workbook(file_name)
        self.sheet_names = self.workbook.sheetnames
        self.output_file_name = out_name_and_ext

    def save(self):
        output_name = "new_" + self.file_name
        if self.output_file_name:
            output_name = self.output_file_name

        self.workbook.save(output_name)

    def replace_links(self):
        for sheet_name in self.sheet_names:
            worksheet = self.workbook[sheet_name]
            for row in worksheet.rows:
                for cell in row:
                    if cell.hyperlink:
                        try:
                            url = cell.hyperlink.target
                            cell.value = url
                        except:
                            pass

        self.save()


if __name__ == "__main__":
    wb = Workbook("HTRs_-_Price_Links_(2).xlsx")
    wb.replace_links()
